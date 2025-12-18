"""
Cutlist Exporter - Export cutlist data to various formats
File: qt6_app/ui_qt/utils/cutlist_exporter.py
"""

from typing import List, Dict, Any
import csv
import json
from datetime import datetime


class CutlistExporter:
    """Handle export to various formats."""
    
    @staticmethod
    def to_csv(pieces: List[Dict[str, Any]], filepath: str):
        """
        Export cutlist to CSV.
        Format: length,quantity,label
        """
        try:
            with open(filepath, 'w', encoding='utf-8', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['length', 'quantity', 'label'])
                
                for piece in pieces:
                    writer.writerow([
                        piece.get('length', 0),
                        piece.get('quantity', 1),
                        piece.get('label', '')
                    ])
        except Exception as e:
            raise Exception(f"Error exporting CSV: {e}")
    
    @staticmethod
    def to_excel(pieces: List[Dict[str, Any]], results: Dict[str, Any], filepath: str):
        """
        Export cutlist + optimization results to Excel.
        Creates two sheets: Cutlist and Optimization Results.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            
            # Sheet 1: Cutlist
            ws1 = wb.active
            ws1.title = "Cutlist"
            
            # Headers
            headers = ['Length (mm)', 'Quantity', 'Label']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(1, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Data
            for row_idx, piece in enumerate(pieces, 2):
                ws1.cell(row_idx, 1, piece.get('length', 0))
                ws1.cell(row_idx, 2, piece.get('quantity', 1))
                ws1.cell(row_idx, 3, piece.get('label', ''))
            
            # Auto-adjust column widths
            for col in ws1.columns:
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws1.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
            
            # Sheet 2: Optimization Results (if provided)
            if results:
                ws2 = wb.create_sheet("Optimization Results")
                
                row = 1
                ws2.cell(row, 1, "Optimization Summary")
                ws2.cell(row, 1).font = Font(bold=True, size=14)
                row += 2
                
                ws2.cell(row, 1, "Bars used:")
                ws2.cell(row, 2, results.get('bars_used', 0))
                row += 1
                
                ws2.cell(row, 1, "Total waste (mm):")
                ws2.cell(row, 2, results.get('total_waste', 0))
                row += 1
                
                ws2.cell(row, 1, "Efficiency (%):")
                ws2.cell(row, 2, results.get('efficiency', 0))
                row += 2
                
                # Cutting plan
                ws2.cell(row, 1, "Cutting Plan")
                ws2.cell(row, 1).font = Font(bold=True, size=12)
                row += 1
                
                for bar_idx, bar in enumerate(results.get('bars', []), 1):
                    ws2.cell(row, 1, f"Bar {bar_idx}:")
                    ws2.cell(row, 1).font = Font(bold=True)
                    row += 1
                    
                    for piece in bar.get('pieces', []):
                        ws2.cell(row, 2, f"{piece.get('length', 0)} mm")
                        ws2.cell(row, 3, piece.get('label', ''))
                        row += 1
                    
                    ws2.cell(row, 2, f"Waste: {bar.get('waste', 0)} mm")
                    ws2.cell(row, 2).font = Font(italic=True)
                    row += 2
            
            wb.save(filepath)
            
        except ImportError:
            raise Exception("openpyxl not installed. Install with: pip install openpyxl")
        except Exception as e:
            raise Exception(f"Error exporting Excel: {e}")
    
    @staticmethod
    def to_pdf(results: Dict[str, Any], filepath: str):
        """
        Export visual cut plan to PDF.
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.pdfgen import canvas
            from reportlab.lib import colors
            
            c = canvas.Canvas(filepath, pagesize=A4)
            width, height = A4
            
            # Title
            c.setFont("Helvetica-Bold", 16)
            c.drawString(50, height - 50, "Cut Plan - Optimization Results")
            
            # Summary
            c.setFont("Helvetica", 12)
            y = height - 80
            c.drawString(50, y, f"Bars used: {results.get('bars_used', 0)}")
            y -= 20
            c.drawString(50, y, f"Total waste: {results.get('total_waste', 0)} mm")
            y -= 20
            c.drawString(50, y, f"Efficiency: {results.get('efficiency', 0):.1f}%")
            y -= 40
            
            # Cutting plan
            c.setFont("Helvetica-Bold", 14)
            c.drawString(50, y, "Cutting Plan:")
            y -= 30
            
            c.setFont("Helvetica", 10)
            for bar_idx, bar in enumerate(results.get('bars', []), 1):
                if y < 100:
                    c.showPage()
                    y = height - 50
                
                c.setFont("Helvetica-Bold", 12)
                c.drawString(50, y, f"Bar {bar_idx} ({results.get('stock_length', 0)} mm):")
                y -= 20
                
                c.setFont("Helvetica", 10)
                for piece in bar.get('pieces', []):
                    label = piece.get('label', '')
                    text = f"  • {piece.get('length', 0)} mm"
                    if label:
                        text += f" [{label}]"
                    c.drawString(70, y, text)
                    y -= 15
                
                c.setFont("Helvetica-Oblique", 10)
                c.drawString(70, y, f"Waste: {bar.get('waste', 0)} mm")
                y -= 30
            
            c.save()
            
        except ImportError:
            raise Exception("reportlab not installed. Install with: pip install reportlab")
        except Exception as e:
            raise Exception(f"Error exporting PDF: {e}")
    
    @staticmethod
    def to_json(project: Dict[str, Any], filepath: str):
        """
        Export project to JSON file.
        """
        try:
            # Add timestamp if not present
            if 'modified_at' not in project:
                project['modified_at'] = datetime.now().isoformat()
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(project, f, indent=2, ensure_ascii=False)
                
        except Exception as e:
            raise Exception(f"Error exporting JSON: {e}")
